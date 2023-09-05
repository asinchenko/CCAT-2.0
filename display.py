# Display to console and html files config check results
# Use display_results function for options display!

# Colored output for windows
from tkinter import W
from fonts import * 
from ctypes import wstring_at
import time
from openpyxl import load_workbook, Workbook

import colorama
colorama.init()

global wb
wb = Workbook()
global ws 
ws = wb.active

global countlines
countlines = 3

class bcolors:
    WHITE  = '\033[1;37m'
    BLUE   = '\033[1;34m'
    GREEN  = '\033[1;32m'
    YELLOW = '\033[1;33m'
    RED    = '\033[1;31m'
    END    = '\033[0m'



# Options display
# Input:  result dictionary of 1 section, option full name
# Sample:    {'dhcp_snooping': {'active': [0, 'DISABLED', 'Turn it off to prevent spoofing attack']}}
# Output: display colored options with its status
# Sample:    - dhcp_snooping active        [DISABLED]
def display_options(dictionary, full_name, html, no_console_display):
    
    color = ''
    htmlclr = ''
    for key in dictionary:
        # color assigning based on severity
        if type(dictionary[key]) is list:
            if dictionary[key][0]   == 0:
                color   = bcolors.RED
                htmlclr = 'red'
            elif dictionary[key][0] == 1:
                color   = bcolors.YELLOW
                htmlclr = '#ff7f00'
            elif dictionary[key][0] == 2:
                color   = bcolors.GREEN
                htmlclr = 'green'
            elif dictionary[key][0] == 3:
                color   = bcolors.WHITE
                htmlclr = 'black'

            # Print option and status to console if needed
            if not no_console_display:
                print('{0:50} {1:1}'.format(' - '+full_name + key, '['+(color+dictionary[key][1]+bcolors.END)+']'))

            # Print option and status to html file if needed
            if html:
                html.write('<tr><td>' + ' - '+full_name + key + '</td>' + '<td style="font-weight: bold; color: '+htmlclr+';">['+dictionary[key][1]+ ']</td></tr>\n')
                # Try to print 'best practice' to html file
                try:
                    html.write('<tr><td></td>' + '<td>*'+dictionary[key][2]+'</td></tr>\n')
                except IndexError:
                    pass
        else:
            # Go deeper in dictionary structure
            full_name += key + ' '
            display_options(dictionary[key],full_name,html, no_console_display)
            full_name = ''


# Results display
# Input:  result dictionary
# Sample:    {'ip':{...}, 'active_service': {...},...}
# Output: colored separated options display
# Sample:    ip
#          - dhcp_snooping active        [DISABLED]
def display_results(dictionary,html_file, no_console_display):
    font = fonts['display_results']
    alignment = alignments['display_results']
    global wb
    global ws
    global countlines

    ws = wb.active
    try:
        ws['A2'] = html_file[:-4]
    except:
        print('A2 error')
    try:
        ws['B2'] = dictionary['model'][0]
    except:
        print('B2 error')
    try:
        ws['C2'] = dictionary['model'][1]
    except:
        print('C2 error')
    try:
        ws['D2'] = dictionary['serial']
    except:
        print('D2 error')
    try:
        ws['E2'] = dictionary['mac']
    except:
        print('E2 error')
    try:
        ws['F2'] = dictionary['os']
    except:
        print('F2 error')    
    try:
        ws['G2'] = dictionary['ios']
    except:
        print('G2 error')      
        
    if html_file:
        # Create and open new .html file
        with open(html_file,'w') as html:
            html.write('<!doctype html>\n<html>\n<head>\n</head>\n<body>\n<table>')
            for key in dictionary:
                full_name = ''

                # Print field name to console if needed
                if not no_console_display:
                    print('\n', bcolors.BLUE + key + bcolors.END)
                # Print field name to html file
                html.write('<tr><td><font color="blue">' + key + '</font></td></tr>\n')
                # Print options in this field
                if key not in ['model', 'mac', 'serial', 'os', 'ios', 'spanning-tree']:
                    display_options(dictionary[key], full_name, html, no_console_display)
                #html.write('<tr><td>&nbsp;</td></tr>\n')
                
            # Html file ending
            html.write('</table>\n</body>\n</html>')
        for key in dictionary:
            if key not in ['model', 'mac', 'serial', 'os', 'ios', 'spanning-tree']:
                full_name = ''
                ws['B'+str(countlines+1)] = key
                ws['B'+str(countlines+1)].font = font
                ws['B'+str(countlines+1)].alignment = alignment
                countlines = countlines + 1
                # Print options in this field
                display_options_excel(dictionary[key], full_name, no_console_display, html_file)
        table_stylyzer()
        wb.save(html_file[:-4] + 'xlsx')
        wb.remove(ws)
        wb.create_sheet('Sheet')
        countlines = 3
    elif not no_console_display:
        for key in dictionary:
            full_name = ''
            print('\n', bcolors.BLUE + key + bcolors.END)
            display_options(dictionary[key], full_name, html_file, no_console_display, html_file)

def display_options_excel(dictionary, full_name, no_console_display, html_name):
    global countlines
    color = ''
    htmlclr = ''
    font = fonts['display_options_excel'] 
    alignment =  alignments['display_options_excel']
    for key in dictionary:
        # color assigning based on severity
        if key not in ['model', 'mac', 'serial', 'os', 'ios', 'spanning-tree']:
            if type(dictionary[key]) is list:
                if dictionary[key][0]   == 0:
                    color   = bcolors.RED
                    htmlclr = '00FF1100'
                elif dictionary[key][0] == 1:
                    color   = bcolors.YELLOW
                    htmlclr = '00FF0000'
                elif dictionary[key][0] == 2:
                    color   = bcolors.GREEN
                    htmlclr = '00378805'
                elif dictionary[key][0] == 3:
                    color   = bcolors.WHITE
                    htmlclr = '00010101'
                
                # Print option and status to console if needed
                if not no_console_display:
                    print('{0:50} {1:1}'.format(' - '+full_name + key, '['+(color+dictionary[key][1]+bcolors.END)+']'))

                

                # Print option and status to html file if needed
                ws.column_dimensions['B'].width = 20
                ws["C"+str(countlines)] = full_name 
                ws["C"+str(countlines)].font = font
                ws.column_dimensions['C'].width = 12
                ws["D"+str(countlines)] = key
                ws.column_dimensions['D'].width = 20
                ws["D"+str(countlines)].font = font
                ws["E"+str(countlines)] = dictionary[key][1]
                ws["E"+str(countlines)].font = font
                ws["E"+str(countlines)].font = Font(color=htmlclr)
                ws["E"+str(countlines)].alignment = alignment
                ws.column_dimensions['E'].width = 10
                try:
                    ws["F"+str(countlines)] = dictionary[key][2]
                    ws["F"+str(countlines)].font = font
                    ws.column_dimensions['F'].width = 120
                    
                except IndexError:
                    pass
            else:
                # Go deeper in dictionary structure
                full_name += key + ' '
                display_options_excel(dictionary[key],full_name, no_console_display, html_name)
                full_name = ''
        countlines = countlines + 1
    #wb.save(html_name[:-4] + 'xlsx')

def table_stylyzer():
    global wb
    global ws
    global countlines
    empty_counter = 0
    merging_point = 0
    for i in (countlines,len(list(ws.rows))):
        if ws["B"+str(i)].value != None:
            merging_point = i
        elif ws["B"+str(i)].value == None and ws["B"+str(i+1)].value == None:
            empty_counter += 1
        elif ws["B"+str(i)].value == None and ws["B"+str(i+1)].value != None:
            empty_counter += 1
        pass